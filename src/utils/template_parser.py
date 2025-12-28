"""
Ship Template Parser - Reads completed Excel template and loads tank data.
"""

from pathlib import Path
from typing import Dict, List, Optional, Tuple
import pandas as pd

try:
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


class TemplateParseResult:
    """Result of parsing a ship template."""
    
    def __init__(self):
        self.success = False
        self.error_message = ""
        self.tank_ids: List[str] = []
        self.ullage_tables: Dict[str, pd.DataFrame] = {}  # {tank_id: DataFrame}
        self.trim_tables: Dict[str, pd.DataFrame] = {}    # {tank_id: DataFrame}
        self.thermal_tables: Dict[str, pd.DataFrame] = {} # {tank_id: DataFrame with temp_c, corr_factor}


def parse_ship_template(filepath: str) -> TemplateParseResult:
    """
    Parse a completed ship template Excel file.
    
    Args:
        filepath: Path to the Excel file
        
    Returns:
        TemplateParseResult object
    """
    result = TemplateParseResult()
    
    if not OPENPYXL_AVAILABLE:
        result.error_message = "openpyxl not installed"
        return result
    
    try:
        wb = load_workbook(filepath, data_only=True)
        
        # Parse Ullage Tables
        if "ULLAGE_TABLES" in wb.sheetnames:
            _parse_ullage_sheet(wb["ULLAGE_TABLES"], result)
        else:
            result.error_message = "ULLAGE_TABLES sheet not found"
            return result
        
        # Parse Trim Correction (optional)
        if "TRIM_CORRECTION" in wb.sheetnames:
            _parse_trim_sheet(wb["TRIM_CORRECTION"], result)
        
        # Parse Thermal Correction (optional)
        if "THERMAL_CORRECTION" in wb.sheetnames:
            _parse_thermal_sheet(wb["THERMAL_CORRECTION"], result)
        
        result.success = True
        return result
        
    except Exception as e:
        result.error_message = f"Error parsing template: {e}"
        return result


def _parse_ullage_sheet(ws, result: TemplateParseResult):
    """Parse the ULLAGE_TABLES sheet."""
    # Read header row to find tank IDs
    headers = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            headers.append((col, str(header)))
    
    # Parse column pairs (ULLAGE, VOLUME)
    col_idx = 0
    while col_idx < len(headers):
        col_num, header = headers[col_idx]
        
        if "_ULLAGE_mm" in header:
            # Extract tank ID
            tank_id = header.replace("_ULLAGE_mm", "")
            result.tank_ids.append(tank_id)
            
            # Find corresponding volume column
            volume_col = col_num + 1
            
            # Read data rows
            ullage_data = []
            for row in range(2, ws.max_row + 1):
                ullage_val = ws.cell(row=row, column=col_num).value
                volume_val = ws.cell(row=row, column=volume_col).value
                
                if ullage_val is not None and volume_val is not None:
                    try:
                        ullage_data.append({
                            'ullage_mm': int(float(ullage_val)),
                            'volume_m3': float(volume_val)
                        })
                    except (ValueError, TypeError):
                        continue
            
            if ullage_data:
                df = pd.DataFrame(ullage_data)
                # Convert mm to cm for compatibility (divide by 10)
                # The system uses CM internally for many calculations, but templates use MM
                df['ullage_cm'] = df['ullage_mm'] / 10.0
                df = df.sort_values('ullage_cm').reset_index(drop=True)
                result.ullage_tables[tank_id] = df
        
        col_idx += 1


def _parse_trim_sheet(ws, result: TemplateParseResult):
    """Parse the TRIM_CORRECTION sheet."""
    current_tank = None
    
    for row in range(1, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=1).value
        
        if cell_val and isinstance(cell_val, str):
            # Check for tank header
            if cell_val.startswith("Tank:"):
                current_tank = cell_val.replace("Tank:", "").strip()
                continue
            
            # Check for header row
            if cell_val == "Ullage_mm":
                # Dynamically determine trim values from headers
                trim_headers = [] # List of (column_index, trim_value)
                for col_idx in range(2, ws.max_column + 1):
                    header_val = ws.cell(row=row, column=col_idx).value
                    if header_val is not None:
                        try:
                            # Try to extract number from string like "-2.0m" or "+0.5m" or just "-2.0"
                            # We normalize the string by removing 'm' and '+' and then converting to float
                            header_str = str(header_val).lower().replace('m', '').replace('+', '').strip()
                            trim_val = float(header_str)
                            trim_headers.append((col_idx, trim_val))
                        except (ValueError, TypeError):
                            continue
                
                # Read all data rows below until empty
                # We expect rows of: Ullage | Correction1 | Correction2 | ...
                trim_data = []
                data_row = row + 1
                while data_row <= ws.max_row:
                    ullage_val = ws.cell(row=data_row, column=1).value
                    if ullage_val is None:
                        break
                    
                    try:
                        ullage_mm = int(float(ullage_val))
                        
                        for col_idx, trim_val in trim_headers:
                            correction = ws.cell(row=data_row, column=col_idx).value
                            if correction is not None:
                                trim_data.append({
                                    'ullage_cm': ullage_mm / 10.0,
                                    'trim_m': trim_val,
                                    'correction_m3': float(correction)
                                })
                    except (ValueError, TypeError):
                        pass # Skip invalid rows
                    
                    data_row += 1
                
                if current_tank and trim_data:
                    # If we already have data for this tank (from another section?), append it
                    if current_tank in result.trim_tables:
                        existing_df = result.trim_tables[current_tank]
                        new_df = pd.DataFrame(trim_data)
                        result.trim_tables[current_tank] = pd.concat([existing_df, new_df]).drop_duplicates().reset_index(drop=True)
                    else:
                        result.trim_tables[current_tank] = pd.DataFrame(trim_data)


def _parse_thermal_sheet(ws, result: TemplateParseResult):
    """Parse the THERMAL_CORRECTION sheet - same format as Ullage Tables."""
    # Read header row to find tank IDs
    headers = []
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            headers.append((col, str(header)))
    
    # Parse column pairs (TEMP_C, CORR_FACTOR)
    col_idx = 0
    while col_idx < len(headers):
        col_num, header = headers[col_idx]
        
        if "_TEMP_C" in header:
            # Extract tank ID
            tank_id = header.replace("_TEMP_C", "")
            
            # Find corresponding correction factor column
            factor_col = col_num + 1
            
            # Read data rows
            thermal_data = []
            for row in range(2, ws.max_row + 1):
                temp_val = ws.cell(row=row, column=col_num).value
                factor_val = ws.cell(row=row, column=factor_col).value
                
                if temp_val is not None and factor_val is not None:
                    try:
                        # Ensure 6 decimal places
                        factor = float(factor_val)
                        thermal_data.append({
                            'temp_c': int(float(temp_val)),
                            'corr_factor': round(factor, 6)
                        })
                    except (ValueError, TypeError):
                        continue
            
            if thermal_data:
                df = pd.DataFrame(thermal_data)
                df = df.sort_values('temp_c').reset_index(drop=True)
                result.thermal_tables[tank_id] = df
        
        col_idx += 1

