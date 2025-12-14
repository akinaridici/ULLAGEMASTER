"""
Ship Template Generator - Creates blank Excel template for user to fill in tank data.
"""

from pathlib import Path
from typing import List
import os

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Styles
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
INSTRUCTION_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
INSTRUCTION_FONT = Font(italic=True, color="806000", size=10)
INPUT_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def generate_ship_template(
    ship_name: str,
    tank_ids: List[str],
    output_path: str,
    include_thermal: bool = True
) -> bool:
    """
    Generate an Excel template for the user to fill in tank calibration data.
    
    Args:
        ship_name: Name of the ship
        tank_ids: List of tank IDs (e.g., ['1P', '1S', '2P', '2S', ...])
        output_path: Path to save the Excel file
        include_thermal: Whether to include thermal correction sheet
        
    Returns:
        True if successful
    """
    if not OPENPYXL_AVAILABLE:
        print("openpyxl not installed")
        return False
    
    try:
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Create Instructions sheet
        _create_instructions_sheet(wb, ship_name, tank_ids, include_thermal)
        
        # Create Ullage Tables sheet
        _create_ullage_sheet(wb, tank_ids)
        
        # Create Trim Correction sheet
        _create_trim_sheet(wb, tank_ids)
        
        # Create Thermal Correction sheet (optional)
        if include_thermal:
            _create_thermal_sheet(wb, tank_ids)
        
        # Save
        wb.save(output_path)
        return True
        
    except Exception as e:
        print(f"Error generating template: {e}")
        return False


def _create_instructions_sheet(wb: Workbook, ship_name: str, tank_ids: List[str], include_thermal: bool):
    """Create the Instructions sheet."""
    ws = wb.create_sheet("INSTRUCTIONS", 0)
    
    # Title
    ws['A1'] = f"ULLAGE TABLES - {ship_name}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:F1')
    
    # Instructions
    instructions = [
        "",
        "HOW TO FILL THIS TEMPLATE:",
        "",
        "1. ULLAGE TABLES Sheet:",
        "   - Each column pair represents one tank (Ullage in mm, Volume in m³)",
        "   - Enter ullage values as 4-digit integers (e.g., 0000, 0100, 0200...)",
        "   - Ullage is measured from the TOP of the tank in millimeters (mm)",
        "   - Volume is the cargo volume at that ullage in cubic meters (m³)",
        "   - Start from ullage = 0 (full tank) to maximum ullage (empty tank)",
        "",
        "2. TRIM CORRECTION Sheet:",
        "   - Correction factors for different ullage levels and trim values",
        "   - Trim = Draft AFT - Draft FWD (positive = stern down)",
        "   - Enter correction in m³ (positive or negative)",
        "",
    ]
    
    if include_thermal:
        instructions.extend([
            "3. THERMAL CORRECTION Sheet:",
            "   - Each column pair represents one tank (Temp in °C, Correction Factor)",
            "   - Enter temperature values as integers (-10, -9, ... 0, 1, 2, ... 50)",
            "   - Enter correction factor with 6 decimal places (e.g., 1.000120)",
            "   - Correction factor adjusts tank volume for steel expansion",
            "",
        ])
    
    instructions.extend([
        "IMPORTANT NOTES:",
        f"- This template is configured for {len(tank_ids)} tanks: {', '.join(tank_ids)}",
        "- Do NOT change the column headers or sheet names",
        "- Leave cells empty if data is not available",
        "- Save this file and upload it back to UllageMaster",
    ])
    
    for i, text in enumerate(instructions, 3):
        cell = ws.cell(row=i, column=1, value=text)
        if text.startswith("HOW TO") or text.startswith("IMPORTANT"):
            cell.font = Font(bold=True, size=12)
        elif text.startswith("   -"):
            cell.font = INSTRUCTION_FONT
    
    # Auto-width
    ws.column_dimensions['A'].width = 80


def _create_ullage_sheet(wb: Workbook, tank_ids: List[str]):
    """Create the Ullage Tables sheet."""
    ws = wb.create_sheet("ULLAGE_TABLES")
    
    # Header row with tank IDs
    col = 1
    for tank_id in tank_ids:
        # Ullage column header
        cell = ws.cell(row=1, column=col, value=f"{tank_id}_ULLAGE_mm")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Volume column header
        cell = ws.cell(row=1, column=col+1, value=f"{tank_id}_VOLUME_m3")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col+1)].width = 15
        
        # Pre-fill some sample ullage values (row 2 onwards)
        sample_ullages = [0, 100, 200, 300, 400, 500, 1000, 1500, 2000, 3000, 4000, 5000, 
                         6000, 7000, 8000, 9000, 10000, 11000, 12000, 13000, 14000, 15000]
        for row_idx, ullage in enumerate(sample_ullages, 2):
            cell = ws.cell(row=row_idx, column=col, value=ullage)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            cell.number_format = '0'
            
            # Volume cell (empty - user fills)
            cell = ws.cell(row=row_idx, column=col+1)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            cell.number_format = '0.000'
        
        col += 2
    
    # Instruction row at bottom
    ws.cell(row=26, column=1, value="Add more rows as needed. Ullage in mm (4 digits), Volume in m³").font = INSTRUCTION_FONT


def _create_trim_sheet(wb: Workbook, tank_ids: List[str]):
    """Create the Trim Correction sheet."""
    ws = wb.create_sheet("TRIM_CORRECTION")
    
    # Header explanation
    ws['A1'] = "TRIM CORRECTION TABLES"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Enter correction values in m³ for each ullage level and trim combination"
    ws['A2'].font = INSTRUCTION_FONT
    
    # For each tank, create a section
    start_row = 4
    for tank_id in tank_ids:
        # Tank header
        ws.cell(row=start_row, column=1, value=f"Tank: {tank_id}").font = Font(bold=True, size=12)
        
        # Column headers: Ullage, then trim values
        headers = ["Ullage_mm", "-2.0m", "-1.5m", "-1.0m", "-0.5m", "0.0m", "+0.5m", "+1.0m", "+1.5m", "+2.0m"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(col_idx)].width = 10
        
        # Sample ullage values
        sample_ullages = [500, 1000, 2000, 3000, 4000, 5000, 6000, 7000, 8000, 9000, 10000]
        for row_offset, ullage in enumerate(sample_ullages, 2):
            cell = ws.cell(row=start_row+row_offset, column=1, value=ullage)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            
            # Empty cells for trim corrections
            for col_idx in range(2, 11):
                cell = ws.cell(row=start_row+row_offset, column=col_idx)
                cell.fill = INPUT_FILL
                cell.border = THIN_BORDER
                cell.number_format = '0.000'
        
        start_row += 15  # Move to next tank section


def _create_thermal_sheet(wb: Workbook, tank_ids: List[str]):
    """Create the Thermal Correction sheet - same format as Ullage Tables."""
    ws = wb.create_sheet("THERMAL_CORRECTION")
    
    # Header row with tank IDs
    col = 1
    for tank_id in tank_ids:
        # Temp column header
        cell = ws.cell(row=1, column=col, value=f"{tank_id}_TEMP_C")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Correction factor column header
        cell = ws.cell(row=1, column=col+1, value=f"{tank_id}_CORR_FACTOR")
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col+1)].width = 18
        
        # Pre-fill temperature values from -10 to 50°C (every 1 degree)
        sample_temps = list(range(-10, 51))  # -10 to 50
        for row_idx, temp in enumerate(sample_temps, 2):
            cell = ws.cell(row=row_idx, column=col, value=temp)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            cell.number_format = '0'
            
            # Correction factor cell (empty - user fills, 6 decimals)
            cell = ws.cell(row=row_idx, column=col+1)
            cell.fill = INPUT_FILL
            cell.border = THIN_BORDER
            cell.number_format = '0.000000'
        
        col += 2
    
    # Instruction row at bottom
    instruction_row = 2 + len(list(range(-10, 51))) + 1
    ws.cell(row=instruction_row, column=1, 
            value="Enter correction factor with 6 decimal places (e.g., 1.000120). Factor=1.0 means no correction.").font = INSTRUCTION_FONT


def get_template_filename(ship_name: str) -> str:
    """Generate a safe filename for the template."""
    safe_name = ship_name.replace(" ", "_").replace("/", "-")
    return f"{safe_name}_ULLAGE_TEMPLATE.xlsx"
