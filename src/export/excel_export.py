"""
Excel export using openpyxl.
"""

from typing import TYPE_CHECKING

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

if TYPE_CHECKING:
    from ..models.voyage import Voyage


# Color definitions
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") if OPENPYXL_AVAILABLE else None
INPUT_FILL = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid") if OPENPYXL_AVAILABLE else None
WARNING_HIGH_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") if OPENPYXL_AVAILABLE else None
WARNING_LOW_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid") if OPENPYXL_AVAILABLE else None
WARNING_CRITICAL_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid") if OPENPYXL_AVAILABLE else None


def export_to_excel(voyage: 'Voyage', filepath: str) -> bool:
    """
    Export voyage data to Excel file.
    
    Args:
        voyage: Voyage object with all tank readings
        filepath: Output file path (.xlsx)
        
    Returns:
        True if successful
    """
    if not OPENPYXL_AVAILABLE:
        print("openpyxl not installed. Run: pip install openpyxl")
        return False
    
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ullage Report"
        
        # Header styling
        header_font = Font(bold=True, color="FFFFFF")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Voyage info section
        ws['A1'] = "Voyage No"
        ws['B1'] = voyage.voyage_number
        ws['C1'] = "Date"
        ws['D1'] = voyage.date
        
        ws['A2'] = "Port"
        ws['B2'] = voyage.port
        ws['C2'] = "Terminal"
        ws['D2'] = voyage.terminal
        
        ws['A3'] = "V.E.F."
        ws['B3'] = voyage.vef
        
        ws['A4'] = "Draft AFT"
        ws['B4'] = voyage.drafts.aft
        ws['C4'] = "Draft FWD"
        ws['D4'] = voyage.drafts.fwd
        ws['E4'] = "Trim"
        ws['F4'] = voyage.drafts.trim
        
        # Column headers (row 6)
        headers = [
            "Tank", "Parcel", "Grade", "Receiver", "Ullage", "% Fill",
            "TOV", "Trim Corr", "GOV", "Temp", "VCF", "GSV",
            "Dens VAC", "Dens Air", "MT (Air)", "MT (VAC)", "Disc."
        ]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            cell.border = thin_border
        
        # Tank data
        row = 7
        for tank_id, reading in voyage.tank_readings.items():
            data = [
                tank_id,
                reading.parcel,
                reading.grade,
                reading.receiver,
                reading.ullage,
                reading.fill_percent,
                reading.tov,
                reading.trim_correction,
                reading.gov,
                reading.temp_celsius,
                reading.vcf,
                reading.gsv,
                reading.density_vac,
                reading.density_air,
                reading.mt_air,
                reading.mt_vac,
                reading.discrepancy
            ]
            
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = thin_border
                
                # Apply warning colors
                # These match the UI's traffic light system for safety
                if col == 6:  # Fill percent column
                    if reading.warning == "high_high":
                        cell.fill = WARNING_CRITICAL_FILL
                    elif reading.warning == "high":
                        cell.fill = WARNING_HIGH_FILL
                    elif reading.warning == "low":
                        cell.fill = WARNING_LOW_FILL
                
                # Input cells styling
                # Highlight USER-editable fields for clarity
                if col in [5, 6, 10, 13]:  # Ullage, Fill%, Temp, Dens VAC
                    if reading.warning == "normal":
                        cell.fill = INPUT_FILL
            
            row += 1
        
        # Totals row
        ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row, column=12, value=voyage.total_gsv).font = Font(bold=True)
        ws.cell(row=row, column=15, value=voyage.total_mt).font = Font(bold=True)
        
        # Auto-fit columns
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 12
        
        # Save
        wb.save(filepath)
        return True
        
    except Exception as e:
        print(f"Error exporting to Excel: {e}")
        return False
