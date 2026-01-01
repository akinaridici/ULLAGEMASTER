"""
Template Report Export Module.
Copies a user-provided XLSM template, injects grid data, and saves as a new file.
"""

import shutil
from pathlib import Path
from typing import TYPE_CHECKING, Optional

try:
    import openpyxl
    from openpyxl import load_workbook
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

if TYPE_CHECKING:
    from ..models.voyage import Voyage


def get_template_path() -> Path:
    """
    Get the path to the TEMPLATE directory.
    
    Supports:
    - Normal Python execution
    - PyInstaller frozen EXE
    - Network share execution
    """
    import sys
    import os
    
    # For frozen executable: use the directory where EXE is located
    if getattr(sys, 'frozen', False):
        # sys.executable is the path to the EXE file
        app_dir = Path(sys.executable).parent
    else:
        # For normal Python execution: go up from src/export to app root
        app_dir = Path(__file__).parent.parent.parent
    
    # Template folder should be next to the EXE or in app root
    template_path = app_dir / "TEMPLATE" / "TEMPLATE.XLSM"
    
    # Fallback: also check 'template' (lowercase) folder
    if not template_path.exists():
        template_path = app_dir / "template" / "TEMPLATE.XLSM"
    
    return template_path


def export_template_report(voyage: 'Voyage', output_path: str, column_keys: list,
                           draft_aft: float = 0.0, draft_fwd: float = 0.0,
                           template_path: Path = None) -> bool:
    """
    Export grid data to a copy of the XLSM template.
    
    Args:
        voyage: The voyage containing tank readings data.
        output_path: Path where the new file will be saved.
        column_keys: List of column keys matching the grid order.
        draft_aft: Draft AFT value in meters.
        draft_fwd: Draft FWD value in meters.
        template_path: Optional path to template file. If None, uses default location.
        
    Returns:
        True if successful, False otherwise.
    """
    if not OPENPYXL_AVAILABLE:
        print("openpyxl not available. Install with: pip install openpyxl")
        return False
    
    # Use provided template_path or get default
    if template_path is None:
        template_path = get_template_path()
    
    if not template_path.exists():
        print(f"Template not found: {template_path}")
        return False
    
    try:
        # Copy template to output path
        shutil.copy2(str(template_path), output_path)
        
        # Open the copy (keep_vba=True preserves macros)
        wb = load_workbook(output_path, keep_vba=True)
        
        # ============ DATA Sheet ============
        if "DATA" not in wb.sheetnames:
            print("DATA sheet not found in template. Creating it.")
            wb.create_sheet("DATA")
        
        ws = wb["DATA"]
        
        # Write headers (row 1)
        for col_idx, key in enumerate(column_keys, start=1):
            ws.cell(row=1, column=col_idx, value=key.upper())
        
        # Write DRAFT values in V1 and W1
        ws.cell(row=1, column=22, value=draft_aft)  # Column V = 22
        ws.cell(row=1, column=23, value=draft_fwd)  # Column W = 23
        
        # Write data (row 2 onwards)
        row_idx = 2
        for tank_id, reading in voyage.tank_readings.items():
            for col_idx, key in enumerate(column_keys, start=1):
                value = _get_reading_value(reading, key, voyage)
                ws.cell(row=row_idx, column=col_idx, value=value)
            row_idx += 1
        
        # ============ DATA_PARCEL Sheet ============
        if "DATA_PARCEL" not in wb.sheetnames:
            wb.create_sheet("DATA_PARCEL")
        
        ws_parcel = wb["DATA_PARCEL"]
        
        # Headers for parcel summary
        parcel_headers = ["PARCEL_NO", "GRADE", "RECEIVER", "VAC_DENSITY", "TOV", "GOV", "MT_VAC", "MT_AIR", "COLOR"]
        for col_idx, header in enumerate(parcel_headers, start=1):
            ws_parcel.cell(row=1, column=col_idx, value=header)
        
        # Aggregate data by parcel
        parcel_totals = {}  # {parcel_id: {grade, receiver, tov, gov, mt_vac, mt_air, color}}
        
        for reading in voyage.tank_readings.values():
            pid = reading.parcel_id
            if not pid:
                continue
            
            if pid not in parcel_totals:
                # Get parcel info
                if pid == "0":  # SLOP
                    grade = "SLOP"
                    receiver = ""
                    color = "#9CA3AF"  # Default gray for SLOP
                    density_vac = 0.0
                else:
                    grade = ""
                    receiver = ""
                    color = ""
                    density_vac = 0.0
                    for p in voyage.parcels:
                        if p.id == pid:
                            grade = p.name
                            receiver = p.receiver
                            color = getattr(p, 'color', '')
                            density_vac = getattr(p, 'density_vac', 0.0)
                            break
                
                parcel_totals[pid] = {
                    'grade': grade,
                    'receiver': receiver,
                    'density_vac': density_vac,
                    'tov': 0.0,
                    'gov': 0.0,
                    'mt_vac': 0.0,
                    'mt_air': 0.0,
                    'color': color
                }
            
            # Add values
            parcel_totals[pid]['tov'] += reading.tov or 0.0
            parcel_totals[pid]['gov'] += reading.gov or 0.0
            parcel_totals[pid]['mt_vac'] += reading.mt_vac or 0.0
            parcel_totals[pid]['mt_air'] += reading.mt_air or 0.0
        
        # Write parcel summary data
        row_idx = 2
        for pid, data in parcel_totals.items():
            ws_parcel.cell(row=row_idx, column=1, value=pid)
            ws_parcel.cell(row=row_idx, column=2, value=data['grade'])
            ws_parcel.cell(row=row_idx, column=3, value=data['receiver'])
            ws_parcel.cell(row=row_idx, column=4, value=data['density_vac'])
            ws_parcel.cell(row=row_idx, column=5, value=data['tov'])
            ws_parcel.cell(row=row_idx, column=6, value=data['gov'])
            ws_parcel.cell(row=row_idx, column=7, value=data['mt_vac'])
            ws_parcel.cell(row=row_idx, column=8, value=data['mt_air'])
            ws_parcel.cell(row=row_idx, column=9, value=data['color'])
            row_idx += 1
        
        # ============ DATA_VOYAGE Sheet ============
        if "DATA_VOYAGE" not in wb.sheetnames:
            wb.create_sheet("DATA_VOYAGE")
        
        ws_voyage = wb["DATA_VOYAGE"]
        
        # Headers for voyage info (Column A = Field Name, Column B = Value)
        voyage_data = [
            ("LOADING_PORT", voyage.port),
            ("LOADING_TERMINAL", voyage.terminal),
            ("VOYAGE_NO", voyage.voyage_number),
            ("DATE", voyage.date),
            ("VEF", voyage.vef),
            ("DRAFT_AFT", draft_aft),
            ("DRAFT_FWD", draft_fwd),
            ("CHIEF_OFFICER", voyage.chief_officer),
            ("MASTER", voyage.master),
        ]
        
        for row_idx, (field, value) in enumerate(voyage_data, start=1):
            ws_voyage.cell(row=row_idx, column=1, value=field)
            ws_voyage.cell(row=row_idx, column=2, value=value)
        
        wb.save(output_path)
        wb.close()
        
        return True
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Error exporting template report: {e}")
        return False


def _get_reading_value(reading, key: str, voyage: 'Voyage'):
    """Get a value from reading by key, with special handling for some fields."""
    
    # Special cases first
    if key == "tank_id":
        return reading.tank_id
    elif key == "parcel":
        return reading.parcel_id or ""
    elif key == "grade":
        # Get grade from parcel
        if reading.parcel_id and reading.parcel_id != "0":  # Not SLOP
            for p in voyage.parcels:
                if p.id == reading.parcel_id:
                    return p.name
        return "SLOP" if reading.parcel_id == "0" else ""
    elif key == "receiver":
        # Get receiver from parcel
        if reading.parcel_id and reading.parcel_id != "0":  # Not SLOP
            for p in voyage.parcels:
                if p.id == reading.parcel_id:
                    return p.receiver
        return ""
    elif key == "receiver_tank":
        return reading.tank_id  # Default to tank_id
    elif key == "temp":
        return reading.temp_celsius
    elif key == "ullage":
        return reading.ullage
    elif key == "fill_percent":
        return reading.fill_percent
    elif key == "trim_corr":
        return reading.trim_correction
    elif key == "corrected_ullage":
        return reading.corrected_ullage
    elif key == "therm_corr":
        return reading.therm_corr
    elif key == "density_vac":
        return reading.density_vac
    elif key == "density_air":
        return reading.density_air
    elif key == "mt_vac":
        return reading.mt_vac
    
    # Direct attribute access for remaining keys (tov, gov, vcf, gsv, mt_air, etc.)
    if hasattr(reading, key):
        return getattr(reading, key, None)
    
    return None


