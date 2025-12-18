import sys
from pathlib import Path
import pandas as pd
from openpyxl import Workbook

# Add src to path
src_path = Path("d:/ULLAGEMASTER/src")
if str(src_path) not in sys.path:
    sys.path.insert(0, str(src_path))

from utils.template_parser import parse_ship_template

def create_test_template(filename):
    wb = Workbook()
    
    # Ullage sheet (required)
    ws_ullage = wb.create_sheet("ULLAGE_TABLES")
    ws_ullage.cell(row=1, column=1, value="1P_ULLAGE_mm")
    ws_ullage.cell(row=1, column=2, value="1P_VOLUME_m3")
    ws_ullage.cell(row=2, column=1, value=1000)
    ws_ullage.cell(row=2, column=2, value=100.0)
    
    # Trim sheet with non-standard ranges
    ws_trim = wb.create_sheet("TRIM_CORRECTION")
    ws_trim.cell(row=1, column=1, value="Tank: 1P")
    ws_trim.cell(row=2, column=1, value="Ullage_mm")
    
    # Custom headers: -3.0m, -1.0m, 0.5m, 1.0m
    ws_trim.cell(row=2, column=2, value="-3.0m")
    ws_trim.cell(row=2, column=3, value="-1.0m")
    ws_trim.cell(row=2, column=4, value="0.5m")
    ws_trim.cell(row=2, column=5, value="1.0m")
    
    # Data row 1
    ws_trim.cell(row=3, column=1, value=500)
    ws_trim.cell(row=3, column=2, value=0.1)
    ws_trim.cell(row=3, column=3, value=0.2)
    ws_trim.cell(row=3, column=4, value=0.3)
    ws_trim.cell(row=3, column=5, value=0.4)
    
    # Data row 2 (testing row limit removal)
    ws_trim.cell(row=4, column=1, value=1000)
    ws_trim.cell(row=4, column=2, value=1.1)
    
    wb.save(filename)
    print(f"Test template created: {filename}")

def verify():
    test_file = "test_trim_dynamic.xlsx"
    create_test_template(test_file)
    
    result = parse_ship_template(test_file)
    
    if not result.success:
        print(f"Error parsing: {result.error_message}")
        return

    print("Successfully parsed template!")
    if "1P" in result.trim_tables:
        df = result.trim_tables["1P"]
        print("\nTrim Data for Tank 1P:")
        print(df)
        
        expected_trims = [-3.0, -1.0, 0.5, 1.0]
        actual_trims = sorted(df['trim_m'].unique().tolist())
        
        print(f"\nExpected trims: {expected_trims}")
        print(f"Actual trims:   {actual_trims}")
        
        if sorted(expected_trims) == actual_trims:
            print("\nSUCCESS: Dynamic trim parsing verified!")
        else:
            print("\nFAILURE: Trim mismatch!")
            
        if len(df[df['ullage_cm'] == 100.0]) > 0:
             print("SUCCESS: Row limit removal verified (found row with 1000mm/100cm ullage)!")
    else:
        print("\nFAILURE: No trim data found for Tank 1P")

if __name__ == "__main__":
    verify()
